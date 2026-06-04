import openpyxl
from openpyxl.styles import Font
from tkinter import filedialog
import os
from db import get_connection
from utils.helpers import load_config, parse_date

def export_to_excel(search_term="", filter_abuse="", filter_status="", filter_year="", filter_month=""):
    config = load_config()
    filename = "VAWC_Records.xlsx"
    path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
    if not path:
        return
    
    connection = None
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VAWC Logs"

        # Add title
        ws.cell(row=1, column=1).value = f"REPUBLIC OF THE PHILIPPINES - {config['lgu_name'].upper()} - VAWC RECORDS"
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        headers = ["VAWC No", "Date", "Client Name", "Age", "Contact", "Birthdate", "Address", "Type of Abuse", "Case Status", "Respondent", "Assigned To", "Settled/Follow-up Date", "Remarks", "Referred To"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="1a2a4a", end_color="1a2a4a", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        connection = get_connection()
        cursor = connection.cursor()
        
        query = "SELECT vawc_no, date, client_name, age, contact, birthdate, address, type_of_abuse, case_status, name_of_respondent, assigned_to, follow_up_date, updated_at, remarks, referred_to FROM vawc_logs WHERE is_deleted = 0"
        params = []
        
        if search_term:
            query += " AND (client_name LIKE ? OR vawc_no LIKE ? OR name_of_respondent LIKE ?)"
            term = f"%{search_term}%"
            params.extend([term, term, term])
        if filter_abuse and filter_abuse != "Type of Abuse":
            query += " AND type_of_abuse LIKE ?"
            params.append(f"%{filter_abuse}%")
        if filter_status and filter_status != "Status":
            query += " AND case_status = ?"
            params.append(filter_status)
        if filter_year and filter_year != "Year":
            query += " AND strftime('%Y', date) = ?"
            params.append(filter_year)
        if filter_month and filter_month != "Month":
            query += " AND strftime('%m', date) = ?"
            params.append(filter_month)

        query += " ORDER BY date DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        for row_num, row in enumerate(rows, 3):
            settled_or_follow_up = ""
            try:
                if str(row[8]).strip().lower() == 'settled':
                    settled_or_follow_up = parse_date(row[12])
                else:
                    settled_or_follow_up = parse_date(row[11])
                if settled_or_follow_up:
                    settled_or_follow_up = settled_or_follow_up.strftime("%m/%d/%Y")
                else:
                    settled_or_follow_up = ""
            except Exception:
                settled_or_follow_up = ""

            row_values = [
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
                row[6],
                row[7],
                row[8],
                row[9],
                row[10],
                settled_or_follow_up,
                row[13],
                row[14]
            ]

            for col_num, cell_value in enumerate(row_values, 1):
                ws.cell(row=row_num, column=col_num).value = str(cell_value) if cell_value else ""

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) # Cap width at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(path)
        from tkinter import messagebox
        messagebox.showinfo("Success", "Excel export completed successfully.")
    except Exception as e:
        from tkinter import messagebox
        messagebox.showerror("Export Error", f"Failed to export to Excel: {str(e)}")
    finally:
        if connection:
            connection.close()